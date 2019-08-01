
def setup():
    import os
    import openpyxl as xl
    from datetime import datetime
    from calendar import monthrange

    current_year=datetime.now().year
    
    for class_ in range(1,11):
        os.makedirs(os.path.join("Database","Students","Images",str(class_)))
        
    os.makedirs(os.path.join("Database","Teachers","Images"))
    os.makedirs(os.path.join("Database","Trainedfile"))
    xl.Workbook().save(os.path.join("Database","teachers_and_students.xlsx"))
    open(os.path.join("Database","unique_ids_list.txt"),"w").close()

    ##Creating and setting up requried excel files
    book1=xl.Workbook()
    book2=xl.Workbook()
    book1.remove(book1.active)
    book2.remove(book2.active)
    months=["January","February","March","April","May","June","July","August","September","October","November","December"]
    
    for sheet_no in range(12):
        book1.create_sheet(months[sheet_no],sheet_no)
        book2.create_sheet(months[sheet_no],sheet_no)
        book1_sheet=book1[months[sheet_no]]
        book2_sheet=book2[months[sheet_no]]
        book1_sheet.column_dimensions['A'].width=18
        book2_sheet.column_dimensions['A'].width=18
        book1_sheet['A1']="ADMISSION NO"
        book2_sheet['A1']="ID NO"
        book1_sheet.column_dimensions['B'].width=20
        book2_sheet.column_dimensions['B'].width=20
        book1_sheet['B1']="NAMES"
        book2_sheet['B1']="NAMES"

        for day in range(monthrange(current_year,sheet_no+1)[1]):
            book1_sheet.cell(row=1,column=3+day).value=day+1
            book2_sheet.cell(row=1,column=3+day).value=day+1
        
    book1.save(os.path.join("Database","Students","Attendence.xlsx"))
    book2.save(os.path.join("Database","Teachers","Attendence.xlsx"))

if __name__=="__main__":
    setup()
